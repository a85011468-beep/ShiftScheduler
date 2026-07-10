import re

from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton, 
                               QLabel, QDateEdit, QMessageBox, QFileDialog, QTableWidget, 
                               QTableWidgetItem, QHeaderView, QSpinBox, QAbstractSpinBox,
                               QComboBox,QProgressDialog)

from PySide6.QtCore import Qt, QDate, QSettings, QThread, Signal
from PySide6.QtGui import QColor, QFont
from datetime import datetime
import pandas as pd
from engine.solver import ScheduleEngine
from config.settings import OFF_SHIFTS, ALL_STATES, SHIFT_DEMANDS
from database.data_importer import DataImporter
from ui.stats_dialog import StatsDialog
from ui.debug_dialog import DebugDialog
from ui.run_config_dialog import RunConfigDialog


class CustomSpinBox(QSpinBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.setFocusPolicy(Qt.StrongFocus) 

    def focusInEvent(self, event):
        super().focusInEvent(event)
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, self.lineEdit().deselect)

class ExportPreviewDialog(QDialog):
    def __init__(self, db_manager, start_date, end_date, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"預覽與修改班表 ({start_date} 至 {end_date})")
        self.resize(1000, 600)
        self.db = db_manager
        self.start_date = start_date
        self.end_date = end_date
        self.pivot_df = None
        
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        info_label = QLabel("💡 請在此處確認或直接修改班表。修改完成後點擊「確認並匯出」。\n⚠️ 注意：此處的任何修改將會同步更新至系統資料庫！")
        info_label.setStyleSheet("color: #1565C0; font-weight: bold; font-size: 14px;")
        layout.addWidget(info_label)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { font-size: 13px; }
            QTableWidget::item:focus { background-color: #E3F2FD; color: black; font-weight: bold; }
        """)
        layout.addWidget(self.table)
        
        btn_layout = QHBoxLayout()
        self.btn_export = QPushButton("💾 確認並匯出")
        self.btn_export.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; height: 35px; font-size: 14px;")
        self.btn_export.clicked.connect(self.on_confirm)
        
        self.btn_cancel = QPushButton("取消匯出")
        self.btn_cancel.setStyleSheet("height: 35px; font-size: 14px;")
        self.btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_cancel)
        btn_layout.addWidget(self.btn_export)
        layout.addLayout(btn_layout)

    def load_data(self):
        schedules = self.db.get_schedule_by_date_range(self.start_date, self.end_date)
        if not schedules:
            return

        employees = self.db.get_all_active_employees()
        level_dict = {str(e['emp_id']).strip(): str(e['job_level']).strip() for e in employees}
        name_dict = {str(e['emp_id']).strip(): e['name'] for e in employees}

        df = pd.DataFrame(schedules)
        self.pivot_df = df.pivot(index='emp_id', columns='date', values='shift_code').fillna('')
        self.pivot_df.index = self.pivot_df.index.astype(str).str.strip()

        # 排序邏輯 (依職級)
        def sort_key(emp_id):
            level = level_dict.get(emp_id, 'Normal')
            if level == 'Chief': return (0, emp_id)
            elif level == 'M': return (1, emp_id)
            else: return (2, emp_id)

        sorted_emp_ids = sorted(self.pivot_df.index, key=sort_key)
        self.pivot_df = self.pivot_df.reindex(sorted_emp_ids)

        # 設定表格
        self.table.setRowCount(len(self.pivot_df))
        self.table.setColumnCount(len(self.pivot_df.columns))
        self.table.setHorizontalHeaderLabels(self.pivot_df.columns)
        
        y_labels = [f"{emp_id} {name_dict.get(emp_id, '')}" for emp_id in self.pivot_df.index]
        self.table.setVerticalHeaderLabels(y_labels)

        # 填入資料並設定週日樣式
        for row_idx, emp_id in enumerate(self.pivot_df.index):
            for col_idx, date in enumerate(self.pivot_df.columns):
                val = str(self.pivot_df.at[emp_id, date]).strip()
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                
                try:
                    dt = datetime.strptime(date, '%Y-%m-%d')
                    if dt.weekday() == 6: # 週日微紅底色提示
                        item.setBackground(QColor("#FCE4EC"))
                except ValueError:
                    pass
                
                self.table.setItem(row_idx, col_idx, item)

    def on_confirm(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # 遍歷網格，讀取使用者的修改並更新 DataFrame 與資料庫
        for row_idx in range(self.table.rowCount()):
            emp_id = str(self.pivot_df.index[row_idx])
            for col_idx in range(self.table.columnCount()):
                date = str(self.pivot_df.columns[col_idx])
                
                item = self.table.item(row_idx, col_idx)
                new_val = item.text().strip() if item else ""
                old_val = str(self.pivot_df.at[emp_id, date]).strip()
                
                # 如果有變更，才進行資料庫寫入
                if new_val != old_val:
                    self.pivot_df.at[emp_id, date] = new_val
                    
                    cursor.execute("SELECT 1 FROM schedule WHERE emp_id=? AND date=?", (emp_id, date))
                    if cursor.fetchone():
                        cursor.execute("UPDATE schedule SET shift_code=? WHERE emp_id=? AND date=?", (new_val, emp_id, date))
                    else:
                        cursor.execute("INSERT INTO schedule (emp_id, date, shift_code, is_locked) VALUES (?, ?, ?, 0)", (emp_id, date, new_val))
                        
        conn.commit()
        conn.close()
        
        self.accept()
class EngineWorker(QThread):
    finished_signal = Signal(bool, str)
    
    def __init__(self, engine, start_date, end_date, leave_quotas, rule_config, parent=None):
        super().__init__(parent)
        self.engine = engine
        self.start_date = start_date
        self.end_date = end_date
        self.leave_quotas = leave_quotas
        self.rule_config = rule_config
        
    def run(self):
        success, message = self.engine.run_scheduler(
            self.start_date, 
            self.end_date, 
            self.leave_quotas, 
            rule_config=self.rule_config
        )
        # 運算結束後，發送訊號通知主畫面
        self.finished_signal.emit(success, message)

class SchedulerDialog(QDialog):
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🚀 排班與預覽匯出中心")
        self.resize(1100, 700) 
        self.db = db_manager
        
        self.settings = QSettings("MyUnit", "ShiftScheduler")
        
        self.leave_spinboxes = {} 

        self.setup_ui()
        self.refresh_table()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # === 頂部控制區 (第一排：主運算) ===
        ctrl_layout = QHBoxLayout()

        self.btn_check_manpower = QPushButton("🧮 試算人力配額")
        self.btn_check_manpower.setStyleSheet("background-color: #607D8B; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_check_manpower.clicked.connect(self.on_check_manpower_clicked)
        ctrl_layout.addWidget(self.btn_check_manpower)
        
        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        saved_start = self.settings.value("scheduler_start_date", "")
        if saved_start:
            self.date_start.setDate(QDate.fromString(saved_start, "yyyy-MM-dd"))
        else:
            self.date_start.setDate(QDate.currentDate().addDays(-QDate.currentDate().day() + 1))
        self.date_start.dateChanged.connect(self.refresh_table)
        ctrl_layout.addWidget(QLabel("🎯 主運算區間："))
        ctrl_layout.addWidget(self.date_start)

        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        saved_end = self.settings.value("scheduler_end_date", "")
        if saved_end:
            self.date_end.setDate(QDate.fromString(saved_end, "yyyy-MM-dd"))
        else:
            self.date_end.setDate(QDate.currentDate().addDays(self.date_start.date().daysInMonth() - QDate.currentDate().day()))
        self.date_end.dateChanged.connect(self.refresh_table)
        ctrl_layout.addWidget(QLabel("至"))
        ctrl_layout.addWidget(self.date_end)

        self.btn_run = QPushButton("⚡ 執行智能排班")
        self.btn_run.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_run.clicked.connect(self.on_run_engine_clicked)
        ctrl_layout.addWidget(self.btn_run)

        self.btn_debug_run = QPushButton("🐛 Debug 勞基法排班")
        self.btn_debug_run.setStyleSheet("background-color: #9C27B0; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_debug_run.clicked.connect(self.on_debug_run_clicked)
        ctrl_layout.addWidget(self.btn_debug_run)

        self.btn_rule_debug = QPushButton("🔍 班表規則診斷")
        self.btn_rule_debug.setStyleSheet("background-color: #00BCD4; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_rule_debug.clicked.connect(self.on_rule_debug_clicked)
        ctrl_layout.addWidget(self.btn_rule_debug)

        # === 頂部控制區 (第二排：夜班參照與輔助功能) ===
        ctrl_layout2 = QHBoxLayout()
        
        self.night_date_start = QDateEdit()
        self.night_date_start.setCalendarPopup(True)
        saved_night_start = self.settings.value("scheduler_night_start", saved_start)
        if saved_night_start:
            self.night_date_start.setDate(QDate.fromString(saved_night_start, "yyyy-MM-dd"))
        self.night_date_start.dateChanged.connect(self.refresh_table)
        
        self.night_date_end = QDateEdit()
        self.night_date_end.setCalendarPopup(True)
        saved_night_end = self.settings.value("scheduler_night_end", saved_end)
        if saved_night_end:
            self.night_date_end.setDate(QDate.fromString(saved_night_end, "yyyy-MM-dd"))
        self.night_date_end.dateChanged.connect(self.refresh_table)

        ctrl_layout2.addWidget(QLabel("🌙 夜班 QSpinBox 參照區間："))
        ctrl_layout2.addWidget(self.night_date_start)
        ctrl_layout2.addWidget(QLabel("至"))
        ctrl_layout2.addWidget(self.night_date_end)
        ctrl_layout2.addStretch() # 推向左側

        self.btn_import_pre = QPushButton("📥 匯入預排 (Excel)")
        self.btn_import_pre.clicked.connect(self.on_import_pre_schedule_clicked)
        ctrl_layout2.addWidget(self.btn_import_pre)
        
        self.btn_clear_pre = QPushButton("🗑️ 清除區間預排")
        self.btn_clear_pre.clicked.connect(self.on_clear_pre_schedule_clicked)
        ctrl_layout2.addWidget(self.btn_clear_pre)

        self.btn_clear = QPushButton("🗑️ 清空區間未鎖定班表")
        self.btn_clear.setStyleSheet("background-color: #795548; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_clear.clicked.connect(self.on_clear_schedule_clicked)
        ctrl_layout2.addWidget(self.btn_clear)

        self.btn_export = QPushButton("💾 滿意並匯出 Excel")
        self.btn_export.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_export.clicked.connect(self.on_export_clicked)
        ctrl_layout2.addWidget(self.btn_export)

        self.btn_stats = QPushButton("📊 週期班別統計")
        self.btn_stats.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold; height: 35px; padding: 0 15px;")
        self.btn_stats.clicked.connect(self.on_stats_clicked)
        ctrl_layout2.addWidget(self.btn_stats)

        layout.addLayout(ctrl_layout)
        layout.addLayout(ctrl_layout2)

        # === 表格區 ===
        layout.addWidget(QLabel("💡 提示：左側可設定引擎代排的 L(特休)/P(事假) 天數總額。\n UI 日期範圍為 [主運算] 與 [夜班參照] 的最大涵蓋範圍，未選定運算的日子引擎不會更動。"))
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setMouseTracking(True) 
        
        self.table.setStyleSheet("""
            QTableWidget {
                alternate-background-color: #f2f2f2;
                font-size: 13px;
            }
            QTableWidget::item:hover {
                background-color: #FFD54F; 
                color: #000000;
                font-size: 15px; 
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #1E90FF;
                color: #FFFFFF;
                font-size: 15px;
                font-weight: bold;
            }
        """)
        layout.addWidget(self.table)
        
        self.refresh_table()
    # 取得主運算區間
    def get_selected_dates(self):
        return self.date_start.date().toString("yyyy-MM-dd"), self.date_end.date().toString("yyyy-MM-dd")
    # 取得夜班參照區間
    def get_night_dates(self):
        return self.night_date_start.date().toString("yyyy-MM-dd"), self.night_date_end.date().toString("yyyy-MM-dd")
    
    def get_ui_dates(self):
        d1 = self.date_start.date()
        d2 = self.date_end.date()
        d3 = self.night_date_start.date()
        d4 = self.night_date_end.date()
        start_date = min(d1, d3).toString("yyyy-MM-dd")
        end_date = max(d2, d4).toString("yyyy-MM-dd")
        return start_date, end_date

    def refresh_table(self):
        self.table.clearContents()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        
        # 💡 [修改] 表格改為抓取全涵蓋區間
        start_date, end_date = self.get_ui_dates()
        
        employees = self.db.get_all_active_employees()
        if not employees:
            return
    
    def on_rule_debug_clicked(self):
        start_date, end_date = self.get_selected_dates()
        dialog = DebugDialog(self.db, start_date, end_date, self)
        dialog.exec()

    def _get_gradient_color(self, val):
        if val <= 8: return QColor(255, 170, 170)
        elif val >= 15: return QColor(170, 255, 170)
        else:
            ratio = (val - 8) / 7.0
            if ratio <= 0.5:
                r = 255
                g = int(170 + (255 - 170) * (ratio / 0.5))
                b = 170
            else:
                r = int(255 - (255 - 170) * ((ratio - 0.5) / 0.5))
                g = 255
                b = 170
            return QColor(r, g, b)

    def refresh_table(self):
        self.table.clearContents()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        
        start_date, end_date = self.get_selected_dates()
        
        employees = self.db.get_all_active_employees()
        if not employees:
            return

        self.leave_widgets = {} 
        self.remaining_labels = {} 
        
        level_dict = {str(e['emp_id']).strip(): str(e['job_level']).strip() for e in employees}
        name_dict = {str(e['emp_id']).strip(): e['name'] for e in employees}
        emp_ids = [str(e['emp_id']).strip() for e in employees]
        
        date_columns = pd.date_range(start=start_date, end=end_date).strftime('%Y-%m-%d').tolist()
        
        schedules = self.db.get_schedule_by_date_range(start_date, end_date)
        
        if schedules:
            df = pd.DataFrame(schedules)
            pivot_df = df.pivot(index='emp_id', columns='date', values='shift_code').fillna('')
            pivot_df.index = pivot_df.index.astype(str).str.strip()
        else:
            pivot_df = pd.DataFrame(index=emp_ids, columns=date_columns).fillna('')

        pivot_df = pivot_df.reindex(index=emp_ids, columns=date_columns, fill_value='')

        def sort_key(emp_id):
            level = level_dict.get(str(emp_id), 'Normal')
            if level == 'Chief': return (0, str(emp_id))
            elif level == 'M': return (1, str(emp_id))
            else: return (2, str(emp_id))

        sorted_emp_ids = sorted(pivot_df.index, key=sort_key)
        pivot_df = pivot_df.reindex(sorted_emp_ids)

        # 欄位定義
        self.control_keys = ['L', 'P', 'r', 'R', '夜', 'O']
        headers = ["特休(L)", "事假(P)", "休息(r)", "例假(R)", "夜班", "加班(O)", "固定", "剩餘(天)"]
        total_days = len(date_columns)

        all_headers = headers + date_columns

        self.stats_rows = [
            ("早M", ["01早M"]), ("早m", ["01早m"]), 
            ("早B1(含c)", ["01早B1", "01早B1c"]), ("早B2(含c)", ["01早B2", "01早B2c"]),
            ("中A", ["01中A"]), 
            ("午M", ["01午M"]), ("午m", ["01午m"]), 
            ("午B1(含c)", ["01午B1", "01午B1c"]), ("午B2(含c)", ["01午B2", "01午B2c"]), 
            ("夜(B1+B2)", ["01夜B1", "01夜B2"])
        ]

        self.table.setRowCount(len(pivot_df) + 1 + len(self.stats_rows))
        self.table.setColumnCount(len(all_headers))

        for col_idx, header_text in enumerate(all_headers):
            item = QTableWidgetItem(header_text)
            try:
                dt = datetime.strptime(header_text, '%Y-%m-%d')
                if dt.weekday() == 6: 
                    item.setBackground(QColor("#AD1457")) 
                    item.setForeground(QColor("#FFFFFF")) 
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
            except ValueError:
                pass 
            self.table.setHorizontalHeaderItem(col_idx, item)
            
        self.table.horizontalHeader().setStyleSheet("QHeaderView::section { padding: 4px; border: 1px solid #ccc; }")
        
        y_labels = ["⚡ 批次套用"] + [f"{emp_id} {name_dict.get(emp_id, '')}" for emp_id in pivot_df.index] + [f"📊 {label}" for label, _ in self.stats_rows]
        self.table.setVerticalHeaderLabels(y_labels)

        # 批次套用工具列
        def make_batch_updater(key):
            def update_all_emps(val):
                for e_id, widgets in self.leave_widgets.items():
                    w = widgets.get(key)
                    if isinstance(w, QSpinBox):
                        w.setValue(val) if val >= w.minimum() else w.setValue(w.minimum())
            return update_all_emps

        for col_idx, key in enumerate(self.control_keys):
            spin = CustomSpinBox()
            spin.setRange(0, total_days)
            spin.setStyleSheet("background-color: #FFF9C4; color: black; font-weight: bold; border: 1px solid #ccc;")
            spin.valueChanged.connect(make_batch_updater(key))
            self.table.setCellWidget(0, col_idx, spin)

        # 第 0 列其餘格子變灰色唯讀
        for col_idx in range(len(all_headers)):
            if not self.table.cellWidget(0, col_idx):
                item = QTableWidgetItem("")
                item.setBackground(QColor("#E0E0E0"))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable) 
                self.table.setItem(0, col_idx, item)

        def make_updater(emp_id, fixed):
            def update():
                widgets = self.leave_widgets[emp_id]
                consumed = fixed
                for k in self.control_keys:
                    w = widgets.get(k)
                    consumed += w.value() if isinstance(w, QSpinBox) else int(w.text())
                
                rem = total_days - consumed
                lbl_rem = self.remaining_labels[emp_id]
                lbl_rem.setText(str(rem))
                lbl_rem.setStyleSheet(
                    "color: #1565C0; font-weight: bold;" if rem >= 0
                    else "background-color: #FFCDD2; color: #B71C1C; font-weight: bold; border: 1px solid #B71C1C;"
                )
            return update

        # 繪製每一列員工
        for row_idx, emp_id in enumerate(pivot_df.index):
            real_row = row_idx + 1 
            emp_data = pivot_df.loc[emp_id]
            managed_states = ['L', 'P', 'r', 'R', '01夜B1', '01夜B2']

            fixed = sum((emp_data != '') & (~emp_data.isin(managed_states)))

            self.leave_widgets[emp_id] = {}
            updater = make_updater(emp_id, fixed)

            def create_label(col_idx, key, val, style="background-color: #EEEEEE; color: #9E9E9E; border: 1px solid #ccc;"):
                lbl = QLabel(str(val))
                lbl.setAlignment(Qt.AlignCenter)
                lbl.setStyleSheet(style)
                self.table.setCellWidget(real_row, col_idx, lbl)
                self.leave_widgets[emp_id][key] = lbl

            def create_spinbox(col_idx, key, existing_val, limit_days):
                spin = CustomSpinBox()
                spin.setRange(existing_val, limit_days) 
                spin.setValue(existing_val)
                spin.setAlignment(Qt.AlignCenter)
                spin.setStyleSheet("background-color: white; color: black; border: 1px solid #ddd;")
                spin.valueChanged.connect(updater) 
                self.table.setCellWidget(real_row, col_idx, spin)
                self.leave_widgets[emp_id][key] = spin

            create_spinbox(0, 'L', sum(emp_data == 'L'), total_days)
            create_spinbox(1, 'P', sum(emp_data == 'P'), total_days)
            create_spinbox(2, 'r', sum(emp_data == 'r'), total_days)
            create_spinbox(3, 'R', sum(emp_data == 'R'), total_days)
            night_count = sum(emp_data.isin(['01夜B1', '01夜B2']))
            create_spinbox(4, '夜', night_count, total_days)
            create_spinbox(5, 'O', 0, total_days)
            
            
            # [修改] 欄位索引向後推移
            create_label(6, '固定(其他)', fixed)

            rem_label = QLabel()
            rem_label.setAlignment(Qt.AlignCenter)
            self.table.setCellWidget(real_row, 7, rem_label) # 索引改為 7
            self.remaining_labels[emp_id] = rem_label

            # 中間日期網格
            for col_offset, date in enumerate(date_columns):
                val = str(pivot_df.at[emp_id, date]).strip()
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable) 
                
                if val: item.setBackground(QColor("#E0E0E0"))
                if val and val not in ALL_STATES:
                    item.setBackground(QColor("#FFCDD2")) 
                    item.setForeground(QColor("#B71C1C")) 
                
                table_col = col_offset + len(headers)
                self.table.setItem(real_row, table_col, item)

            updater()

        # ==========================================
        # 👑 最底下列：多列班別詳細統計與紅綠燈
        # ==========================================
        last_row_start = len(pivot_df) + 1
        
        for row_offset, (label, shift_codes) in enumerate(self.stats_rows):
            current_row = last_row_start + row_offset
            
            for i in range(len(headers)):
                item = QTableWidgetItem("")
                item.setBackground(QColor("#E0E0E0"))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable) 
                self.table.setItem(current_row, i, item)

            for col_idx, date in enumerate(date_columns):
                col_data = pivot_df[date]
                count = sum(col_data.isin(shift_codes))
                
                max_limit = sum(SHIFT_DEMANDS.get(sc, (0, 1))[1] for sc in shift_codes)
                if max_limit == 0: max_limit = 1 
                
                item = QTableWidgetItem(f"{count}")
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable) 
                
                if count == 0:
                    item.setBackground(QColor("#FFCDD2"))
                    item.setForeground(QColor("#B71C1C"))
                elif count >= max_limit:
                    item.setBackground(QColor("#C8E6C9"))
                    item.setForeground(QColor("#1B5E20"))
                else:
                    item.setBackground(QColor("#FFF9C4"))
                    item.setForeground(QColor("#F57F17"))
                    
                font = QFont()
                font.setBold(True)
                font.setPointSize(11)
                item.setFont(font)
                
                self.table.setItem(current_row, col_idx + len(headers), item)

        # 欄位寬度設定
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.table.horizontalHeader().setDefaultSectionSize(75)
        self.table.verticalHeader().setDefaultSectionSize(35)
        
        for i in range(len(headers)):
            self.table.horizontalHeader().resizeSection(i, 60)
          
    def on_run_engine_clicked(self):
        start_date, end_date = self.get_selected_dates()
        night_start, night_end = self.get_night_dates()
        
        config_dialog = RunConfigDialog(self)
        if config_dialog.exec() != QDialog.Accepted:
            return
            
        rule_config = config_dialog.get_config()
        
        leave_quotas = {}
        for emp_id, widgets in self.leave_widgets.items():
            leave_quotas[emp_id] = {}
            for k in self.control_keys:
                w = widgets.get(k)
                if w is not None:
                    try:
                        val = w.value() 
                    except AttributeError:
                        try:
                            val = int(w.text())
                        except (ValueError, TypeError):
                            val = 0
                    leave_quotas[emp_id][k] = val
                else:
                    leave_quotas[emp_id][k] = 0
                    
        self.engine = ScheduleEngine(self.db)
        
        # 建立來回滾動的進度視窗
        self.progress_dialog = QProgressDialog("⚙️ 引擎正在全力運算中，請稍候...\n(若條件嚴苛可能需要數十秒至數分鐘)", "🛑 放棄本次運算", 0, 0, self)
        self.progress_dialog.setWindowTitle("執行智能排班")
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setMinimumDuration(0) 
        self.progress_dialog.canceled.connect(self.cancel_engine) # 綁定取消事件
        
        # 啟動背景執行緒
        self.worker = EngineWorker(self.engine, start_date, end_date, night_start, night_end, leave_quotas, rule_config)
        self.worker.finished_signal.connect(self.on_engine_finished)
        self.worker.start()
        
        # 顯示視窗 (不使用 exec() 避免死鎖)
        self.progress_dialog.show()

    def cancel_engine(self):
        """使用者按下取消按鈕，發送中斷訊號給引擎"""
        if hasattr(self, 'engine'):
            self.engine.cancel()
        
    def on_engine_finished(self, success, message):
        """背景引擎執行完畢後的 Callback"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
            
        if success:
            QMessageBox.information(self, "排班結果", message)
            self.refresh_table() 
        else:
            QMessageBox.warning(self, "排班結果", message)

    def on_debug_run_clicked(self):
        """執行暴力 Debug 排班前，先彈出分級人力診斷報告"""
        start_date, end_date = self.get_selected_dates()
        dates = pd.date_range(start=start_date, end=end_date).tolist()
        num_days = len(dates)

        from config.settings import SHIFT_DEMANDS, MANAGER_ONLY_SHIFTS
        
        daily_min_m = sum(req[0] for s, req in SHIFT_DEMANDS.items() if s in MANAGER_ONLY_SHIFTS)
        daily_max_m = sum(req[1] for s, req in SHIFT_DEMANDS.items() if s in MANAGER_ONLY_SHIFTS)
        daily_min_n = sum(req[0] for s, req in SHIFT_DEMANDS.items() if s not in MANAGER_ONLY_SHIFTS)
        daily_max_n = sum(req[1] for s, req in SHIFT_DEMANDS.items() if s not in MANAGER_ONLY_SHIFTS)
        
        cycle_min_m = daily_min_m * num_days
        cycle_max_m = daily_max_m * num_days
        cycle_min_n = daily_min_n * num_days
        cycle_max_n = daily_max_n * num_days

        active_employees = self.db.get_all_active_employees()
        managers = [e for e in active_employees if str(e.get('job_level', 'Normal')).strip() in ('M', 'Chief', 't')]
        normals = [e for e in active_employees if str(e.get('job_level', 'Normal')).strip() not in ('M', 'Chief', 't')]
        
        num_m = len(managers)
        num_n = len(normals)
        
        slots_m = num_m * num_days
        slots_n = num_n * num_days

        manager_ids = [str(e['emp_id']).strip() for e in managers]
        total_off_m = 0
        total_off_n = 0
        leave_quotas = {}
        off_keys = ['L', 'P', 'r', 'R']  # O(加班) 不算休假
            
        for emp_id, widgets in self.leave_widgets.items():
            eid_str = str(emp_id).strip()
            leave_quotas[eid_str] = {}
            emp_off_days = 0
            for k in self.control_keys:
                w = widgets.get(k)
                val = 0
                if w is not None:
                    try:
                        val = w.value() 
                    except AttributeError:
                        try:
                            val = int(w.text())
                        except (ValueError, TypeError):
                            val = 0
                leave_quotas[eid_str][k] = val
                if k in off_keys:
                    emp_off_days += val
            
            if eid_str in manager_ids:
                total_off_m += emp_off_days
            else:
                total_off_n += emp_off_days

        avail_m = slots_m - total_off_m
        avail_n = slots_n - total_off_n

        msg = (f"📊 【分級人力配額物理試算】\n\n"
               f"📅 區間：{start_date} 至 {end_date} (共 {num_days} 天)\n"
               f"----------------------------------------\n"
               f"👑 主管級人員 (M/Chief/t)：{num_m} 人\n"
               f"🛏️ 主管總休假數：{total_off_m} 天\n"
               f"🎯 主管總需求下限：{cycle_min_m} 人次 | 上限：{cycle_max_m} 人次\n"
               f"💪 主管實際可用人力：{avail_m} 人次\n")
        
        if avail_m < cycle_min_m or avail_m > cycle_max_m:
            msg += "⚠️ 警告：主管總量超出物理班表邊界！\n"
        else:
            msg += "✅ 評估：主管級總量正常。\n"

        msg += (f"\n👥 基層人員 (Normal)：{num_n} 人\n"
               f"🛏️ 基層總休假數：{total_off_n} 天\n"
               f"🎯 基層總需求下限：{cycle_min_n} 人次 | 上限：{cycle_max_n} 人次\n"
               f"💪 基層實際可用人力：{avail_n} 人次\n")
        
        if avail_n < cycle_min_n or avail_n > cycle_max_n:
            msg += "⚠️ 警告：基層總量超出物理班表邊界！\n"
        else:
            msg += "✅ 評估：基層級總量正常。\n"

        msg += "\n----------------------------------------\n"
        msg += "\n是否確認執行【終極暴力 Debug】？\n(⚠️ 警告：此模式將徹底無視七休一、交接班相剋等所有勞基法規則，僅為測試「人數物理極限」而存在！)"

        reply = QMessageBox.question(self, "暴力 Debug (無視法規)", msg, QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            engine = ScheduleEngine(self.db)
            try:
                success, message = engine.run_scheduler(start_date, end_date, leave_quotas, debug_mode=True, rule_config={'strict_quotas': False, 'hard_mid_a': False, 'hard_no_4_off': False})
            except TypeError:
                success, message = engine.run_scheduler(start_date, end_date, leave_quotas)

            if success:
                QMessageBox.information(self, "排班結果", message)
                self.refresh_table() 
            else:
                QMessageBox.warning(self, "排班失敗", message)

    def on_check_manpower_clicked(self):
        """僅顯示當前 QSpinBox 休假與人力統計資訊，並預檢硬規則衝突，不執行排班"""
        start_date, end_date = self.get_selected_dates()
        night_start, night_end = self.get_night_dates()

        dates = pd.date_range(start=start_date, end=end_date).tolist()
        num_days = len(dates)
        eval_dates = [d.strftime('%Y-%m-%d') for d in dates]
        night_dates = pd.date_range(start=night_start, end=night_end).strftime('%Y-%m-%d').tolist()

        overall_start = min(start_date, night_start)
        overall_end = max(end_date, night_end)

        schedules = self.db.get_schedule_by_date_range(eval_dates[0], eval_dates[-1])
        dict_sched = {(s['emp_id'], s['date']): s for s in schedules}
        
        overall_schedules = self.db.get_schedule_by_date_range(overall_start, overall_end)
        dict_sched_overall = {(s['emp_id'], s['date']): s for s in overall_schedules}

        from config.settings import SHIFT_DEMANDS, MANAGER_ONLY_SHIFTS
        
        daily_min_m = sum(req[0] for s, req in SHIFT_DEMANDS.items() if s in MANAGER_ONLY_SHIFTS)
        daily_max_m = sum(req[1] for s, req in SHIFT_DEMANDS.items() if s in MANAGER_ONLY_SHIFTS)
        daily_min_n = sum(req[0] for s, req in SHIFT_DEMANDS.items() if s not in MANAGER_ONLY_SHIFTS)
        daily_max_n = sum(req[1] for s, req in SHIFT_DEMANDS.items() if s not in MANAGER_ONLY_SHIFTS)
        
        # 👇 抓出夜班的每日需求，試算整個區間需要幾個夜班配額
        req_night_per_day = SHIFT_DEMANDS.get('01夜B1', [0,0])[0] + SHIFT_DEMANDS.get('01夜B2', [0,0])[0]
        min_night_needed = req_night_per_day * num_days

        cycle_min_m = daily_min_m * num_days
        cycle_max_m = daily_max_m * num_days
        cycle_min_n = daily_min_n * num_days
        cycle_max_n = daily_max_n * num_days

        active_employees = self.db.get_all_active_employees()
        managers = [e for e in active_employees if str(e.get('job_level', 'Normal')).strip() in ('M', 'Chief', 't')]
        normals = [e for e in active_employees if str(e.get('job_level', 'Normal')).strip() not in ('M', 'Chief', 't')]
        
        num_m = len(managers)
        num_n = len(normals)
        
        slots_m = num_m * num_days
        slots_n = num_n * num_days

        manager_ids = [str(e['emp_id']).strip() for e in managers]
        total_off_m = 0
        total_off_n = 0
        total_night_quota = 0  # 👇 追蹤已經發出去的夜班數量
        leave_quotas = {}
        off_keys = ['L', 'P', 'r', 'R']  # O(加班) 不算休假

        for emp_id, widgets in self.leave_widgets.items():
            eid_str = str(emp_id).strip()
            leave_quotas[eid_str] = {}
            emp_off_days = 0
            for k in self.control_keys:
                w = widgets.get(k)
                val = 0
                if w is not None:
                    try:
                        val = w.value() 
                    except AttributeError:
                        try:
                            val = int(w.text())
                        except (ValueError, TypeError):
                            val = 0
                leave_quotas[eid_str][k] = val
                if k in off_keys:
                    emp_off_days += val
                
                # 👇 累加夜班 QSpinBox 的總和
                if k == '夜':
                    total_night_quota += val
            
            if eid_str in manager_ids:
                total_off_m += emp_off_days
            else:
                total_off_n += emp_off_days

        avail_m = slots_m - total_off_m
        avail_n = slots_n - total_off_n

        # 硬規則安檢探針
        emp_ids = [e['emp_id'] for e in active_employees]
        job_levels = {str(e['emp_id']).strip(): str(e['job_level']).strip() for e in active_employees}
        shift_prefs = {str(e['emp_id']).strip(): str(e.get('shift_pref', 'MIX')).strip() for e in active_employees}
        
        schedules = self.db.get_schedule_by_date_range(eval_dates[0], eval_dates[-1])
        dict_sched = {(s['emp_id'], s['date']): s for s in schedules}

        from datetime import datetime, timedelta
        eval_start_dt = datetime.strptime(eval_dates[0], '%Y-%m-%d')
        history_start = (eval_start_dt - timedelta(days=7)).strftime('%Y-%m-%d')
        history_end = (eval_start_dt - timedelta(days=1)).strftime('%Y-%m-%d')
        history_schedules = self.db.get_schedule_by_date_range(history_start, history_end)
        dict_history = {(s['emp_id'], s['date']): s['shift_code'] for s in history_schedules}
        
        engine = ScheduleEngine(self.db)
        hard_conflicts = engine._run_pre_flight_diagnostics(
            emp_ids, eval_dates, dict_sched, dict_history, job_levels, shift_prefs, leave_quotas, night_dates, dict_sched_overall
        )

        msg = (f"📊 【分級人力配額物理試算】\n\n"
               f"📅 區間：{start_date} 至 {end_date} (共 {num_days} 天)\n"
               f"----------------------------------------\n"
               f"👑 主管級人員 (M/Chief/t)：{num_m} 人\n"
               f"🛏️ 主管總休假數：{total_off_m} 天\n"
               f"🎯 主管總需求下限：{cycle_min_m} 人次 | 上限：{cycle_max_m} 人次\n"
               f"💪 主管實際可用人力：{avail_m} 人次\n")
        
        if avail_m < cycle_min_m:
            msg += "⚠️ 警告：主管可用人力低於專屬班別需求底線！\n"
        elif avail_m > cycle_max_m:
            msg += "⚠️ 警告：主管總可用人力超出專屬班別上限！\n"
        else:
            msg += "✅ 評估：主管級總量正常。\n"

        msg += (f"\n👥 基層人員 (Normal)：{num_n} 人\n"
               f"🛏️ 基層總休假數：{total_off_n} 天\n"
               f"🎯 基層總需求下限：{cycle_min_n} 人次 | 上限：{cycle_max_n} 人次\n"
               f"💪 基層實際可用人力：{avail_n} 人次\n")
        
        if avail_n < cycle_min_n:
            msg += "⚠️ 警告：基層可用人力低於底線需求！\n"
        elif avail_n > cycle_max_n:
            msg += "⚠️ 警告：基層可工作天數超出全體基層班別最大容量！\n"
        else:
            msg += "✅ 評估：基層級總量正常。\n"
            
        # 👇 新增夜班面板提示，讓你一眼看穿還有幾個夜班額度沒發！
        msg += (f"\n----------------------------------------\n"
                f"🌙 【夜班專屬配額試算】\n"
                f"🎯 本區間夜班總需求：{min_night_needed} 人次\n"
                f"📥 QSpinBox 目前共發放：{total_night_quota} 人次\n")
        if total_night_quota < min_night_needed:
            msg += f"❌ 錯誤：夜班配額不足！還差 {min_night_needed - total_night_quota} 個人次。引擎無人可用。\n"
        elif total_night_quota > min_night_needed:
            msg += f"⚠️ 警告：發放過多 (多出 {total_night_quota - min_night_needed} 人次)，將被強迫安插。\n"
        else:
            msg += f"✅ 評估：夜班配額完美吻合需求！\n"

        msg += "\n----------------------------------------\n"
        msg += "🚨 【預排班表與配額死結預檢】\n"
        if not hard_conflicts:
            msg += "✅ 檢查通過：當前預排圖釘與特休配額無硬性法規死結。\n"
        else:
            msg += f"❌ 偵測到 {len(hard_conflicts)} 項硬規則死結 (會引發0.5秒崩潰)：\n\n"
            for err in hard_conflicts[:5]:
                msg += f"  👉 {err}\n"
            if len(hard_conflicts) > 5:
                msg += f"  ...等共 {len(hard_conflicts)} 項衝突。\n"
            msg += "\n💡 提示：請先至主面板解除圖釘鎖定或調整 QSpinBox 配額！"

        QMessageBox.information(self, "分級人力與死結綜合試算 (不排班)", msg)
    def on_clear_schedule_clicked(self):
        """清空選定區間內所有未鎖定的排班資料"""
        start_date, end_date = self.get_selected_dates()
        
        if self.db.is_period_locked(start_date, end_date):
            QMessageBox.critical(self, "🛑 操作被拒", f"指定區間 ({start_date} 至 {end_date}) 已被系統結算鎖定！\n請先至資料庫管理中心解鎖。")
            return

        reply = QMessageBox.question(self, "確認清空班表", 
                                     f"您確定要清空 {start_date} 至 {end_date} 之間的所有自動排班嗎？\n\n💡 注意：這只會抹除引擎排出的班別，您匯入的歷史班表與鎖定的請假資料（📌圖釘）將會受到保護、維持原狀。",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    DELETE FROM schedule 
                    WHERE date >= ? AND date <= ? AND is_locked = 0
                ''', (start_date, end_date))
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, "清空完成", "✅ 未鎖定之班表已成功清除，回復為空白網格。")
                self.refresh_table()
                
            except Exception as e:
                QMessageBox.critical(self, "錯誤", f"刪除過程中發生異常：\n{str(e)}")

    def on_stats_clicked(self):
        start_date, end_date = self.get_selected_dates()
        dialog = StatsDialog(self.db, start_date, end_date, self)
        dialog.exec()

    def on_export_clicked(self):
        start_date, end_date = self.get_selected_dates()
        schedules = self.db.get_schedule_by_date_range(start_date, end_date)
        if not schedules:
            QMessageBox.warning(self, "匯出失敗", "選定區間內沒有排班資料可匯出！")
            return

        # 啟動預覽與修改視窗
        dialog = ExportPreviewDialog(self.db, start_date, end_date, self)
        
        # 若使用者在修改視窗點擊「確認並匯出」
        if dialog.exec() == QDialog.Accepted:
            # 選擇存檔路徑
            file_path, _ = QFileDialog.getSaveFileName(self, "匯出班表", f"排班表_{start_date}_至_{end_date}.xlsx", "Excel Files (*.xlsx)")
            if file_path:
                try:
                    # 直接匯出剛剛視窗中已更新的 pivot_df
                    dialog.pivot_df.to_excel(file_path)

                    import openpyxl
                    from openpyxl.styles import PatternFill
                    
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    c_shift_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
                    
                    for row in ws.iter_rows(min_row=2, min_col=2):
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value.endswith('c'):
                                cell.fill = c_shift_fill
                                
                    wb.save(file_path)
                    QMessageBox.information(self, "匯出成功", f"檔案已儲存至：\n{file_path}\n\n✅ 您所做的班別修改已同步儲存至資料庫中。")
                    
                    # 重新整理主畫面的排班表，使其與剛才的修改結果同步
                    self.refresh_table()
                except Exception as e:
                    QMessageBox.critical(self, "匯出失敗", f"寫入 Excel 時發生錯誤：\n{str(e)}")

    def closeEvent(self, event):
        start_date_str = self.date_start.date().toString("yyyy-MM-dd")
        end_date_str = self.date_end.date().toString("yyyy-MM-dd")
        night_start_str = self.night_date_start.date().toString("yyyy-MM-dd")
        night_end_str = self.night_date_end.date().toString("yyyy-MM-dd")

        self.settings.setValue("scheduler_start_date", start_date_str)
        self.settings.setValue("scheduler_end_date", end_date_str)
        self.settings.setValue("scheduler_night_start", night_start_str)
        self.settings.setValue("scheduler_night_end", night_end_str)
        super().closeEvent(event)

    def on_import_pre_schedule_clicked(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "選擇預排 Excel 檔案", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return
            
        start_date = self.date_start.date().toString("yyyy-MM-dd")
        importer = DataImporter(self.db)
        success, message = importer.import_pre_schedule(file_path, start_date)
        
        if success:
            self.refresh_table()
            QMessageBox.information(self, "匯入成功", message)
        else:
            QMessageBox.critical(self, "匯入失敗", message)

    def on_clear_pre_schedule_clicked(self):
        start_date, end_date = self.get_selected_dates()
        
        reply = QMessageBox.question(
            self, '確認清除', 
            f'⚠️ 確定要清除 {start_date} 到 {end_date} 之間的所有【鎖定預排】嗎？\n(這將把格子還原為空白，交由引擎重新排班)',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE schedule 
                SET is_locked = 0, shift_code = NULL 
                WHERE date >= ? AND date <= ? AND is_locked = 1
            ''', (start_date, end_date))
            
            deleted_count = cursor.rowcount
            conn.commit()
            conn.close()
            
            self.refresh_table()
            QMessageBox.information(self, "清除成功", f"已解除區間內 {deleted_count} 個預排圖釘！")
